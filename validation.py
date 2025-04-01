import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from difflib import get_close_matches

def clean_columns(df):
    df = df.dropna(axis=1, how="all")
    df.columns = (
        df.columns.astype(str)
        .str.upper()
        .str.strip()
        .str.replace(r'\s+', '_', regex=True)
        .str.replace("'", "")
    )
    return df

def load_and_clean_excel(file_path, sheet_name=None, usecols=None, skiprows=None):
    excel_data = pd.read_excel(file_path, sheet_name=sheet_name, usecols=usecols, skiprows=skiprows)
    if isinstance(excel_data, dict):
        sheet_name = sheet_name if sheet_name else list(excel_data.keys())[0]
        df = excel_data[sheet_name]
    else:
        df = excel_data
    return clean_columns(df)

def find_similar_columns(preload_cols, postload_cols, threshold=0.6):
    matched_columns = {}
    for col in preload_cols:
        matches = get_close_matches(col, postload_cols, n=1, cutoff=threshold)
        if matches:
            matched_columns[col] = matches[0]
    return matched_columns

def user_confirm_column_mapping(matched_columns, postload_cols):
    print("\n📌 Suggested Column Matches (Type 'none' to skip validation for a column):")
    confirmed_matches = {}
    for pre_col, suggested_post_col in matched_columns.items():
        print(f"\n➡️ Preload Column: {pre_col}")
        print(f"   Suggested Postload Match: {suggested_post_col}")
        user_input = input("   ✅ Press Enter to confirm, type a different column name, or 'none' to skip: ").strip()
        if user_input.lower() == "none":
            continue  # Skip validation for this column
        elif user_input and user_input in postload_cols:
            confirmed_matches[pre_col] = user_input
        else:
            confirmed_matches[pre_col] = suggested_post_col
    return confirmed_matches

def highlight_differences(preFile, postFile, outputFile):
    preDf = load_and_clean_excel(preFile, sheet_name="Sheet4", usecols="B:X")
    postDf = load_and_clean_excel(postFile)
    matched_columns = find_similar_columns(preDf.columns.tolist(), postDf.columns.tolist(), threshold=0.6)
    print("PreDf: ",preDf.columns.tolist())
    matched_columns = user_confirm_column_mapping(matched_columns, postDf.columns.tolist())
    wb = load_workbook(postFile)
    ws = wb.active
    difference_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    missing_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    for row in range(len(preDf)):
        for pre_col, post_col in matched_columns.items():
            if pre_col not in preDf.columns or post_col not in postDf.columns:
                continue
            pre_value = preDf.at[row, pre_col] if row < len(preDf) else None
            post_value = postDf.at[row, post_col] if row < len(postDf) else None
            excel_row = row + 2
            excel_col = postDf.columns.get_loc(post_col) + 1
            if pre_value != post_value:
                ws.cell(row=excel_row, column=excel_col).fill = difference_fill
            if pd.notna(pre_value) and (pd.isna(post_value) or post_value == ""):
                ws.cell(row=excel_row, column=excel_col).fill = missing_fill
    wb.save(outputFile)
    print(f"✅ Highlighted differences saved in: {outputFile}")

preFile = "MDG Supplier Master Mass Upload Template- 1st Draft.xlsx"
postFile = "ACVS_FD_Supplier_Master_Postload_V2.xlsx"
outputFile = "Highlighted_Postload.xlsx"

highlight_differences(preFile, postFile, outputFile)

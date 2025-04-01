import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

# File paths
preFile = "MDG Supplier Master Mass Upload Template- 1st Draft.xlsx"
postFile = "ACVS_FD_Supplier_Master_Postload_V2.xlsx"
outputFile = "output.xlsx"



def clean_value(value):
    """Clean and standardize value for comparison"""
    if value is None or pd.isna(value):
        return ''
    # Convert to string, remove all leading/trailing spaces
    cleaned = str(value).strip()
    
    # Handle decimal numbers (convert to integer string)
    try:
        if '.' in cleaned:
            cleaned = str(int(float(cleaned)))
        elif cleaned.isdigit():
            cleaned = str(int(cleaned))
    except ValueError:
        pass
    
    return cleaned

def get_base_column_name(column_name):
    """Get the base column name (first word before space or underscore)"""
    return column_name.split('_')[0].split()[0].strip().upper()

# Load Excel files
pre_data = pd.read_excel(preFile, sheet_name="Sheet4")
post_data = pd.read_excel(postFile)



# Find ALTKN columns
pre_altkn_col = next(col for col in pre_data.columns if 'ALTKN' in col.upper())
post_altkn_col = next(col for col in post_data.columns if 'ALTKN' in col.upper())

# print("pre_altkn_col: ",pre_altkn_col)
# print("post_altkn_col: ",post_altkn_col)

column_mapping = {}

for post_col in post_data.columns:
    post_base = get_base_column_name(post_col)
    # print(" post_base: ",post_base," post_col: ",post_col)
    for pre_col in pre_data.columns:
        # print("pre_col ",pre_col, " post_base: ",post_base)
        if get_base_column_name(pre_col) == post_base:
            column_mapping[post_col] = pre_col

            break

# Write post_data to new Excel
post_data.to_excel(outputFile, index=False)
workbook = load_workbook(outputFile)
worksheet = workbook.active

# Define colors
changed_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')  
missing_fill = PatternFill(start_color='E5A78C', end_color='E5A78C', fill_type='solid')  
blank_key_fill = PatternFill(start_color='E6F3FF', end_color='E6F3FF', fill_type='solid')  


pre_dict = {}
for _, row in pre_data.iterrows():
    altkn = clean_value(row[pre_altkn_col])
    if altkn:
        # Store cleaned values
        cleaned_row = {col: clean_value(val) for col, val in row.items()}
        pre_dict[altkn] = cleaned_row
        print(f"Loaded pre-load ALTKN: '{altkn}'")  # Debug print
        
print("pre_dict: ",pre_dict)

# Compare values
for row in range(2, worksheet.max_row + 1):
    # Get ALTKN from post-data
    post_altkn = clean_value(worksheet.cell(row=row, column=post_data.columns.get_loc(post_altkn_col) + 1).value)
    print(f"\nProcessing post-load ALTKN: '{post_altkn}'")  # Debug print
    
    # If ALTKN is blank, highlight entire row in blue
    if not post_altkn:
        for col_idx in range(1, worksheet.max_column + 1):
            cell = worksheet.cell(row=row, column=col_idx)
            cell.fill = blank_key_fill
        continue
    
    # If ALTKN exists in pre-data, compare all other fields
    if post_altkn in pre_dict:
        print(f"Found match for ALTKN: {post_altkn}")  # Debug print
        for col_idx, post_col_name in enumerate(post_data.columns, 1):
            if post_col_name != post_altkn_col:  # Skip ALTKN column
                cell = worksheet.cell(row=row, column=col_idx)
                post_value = clean_value(cell.value)
                
                # Get corresponding pre-load column name
                pre_col_name = column_mapping.get(post_col_name)
                if pre_col_name:
                    pre_value = clean_value(pre_dict[post_altkn].get(pre_col_name, ''))
                    
                    # Compare and highlight differences
                    if post_value != pre_value:
                        if not post_value and pre_value:
                            print(f"Missing value in {post_col_name}: pre='{pre_value}', post='{post_value}'")
                            cell.fill = missing_fill
                        else:
                            print(f"Changed value in {post_col_name}: pre='{pre_value}', post='{post_value}'")
                            cell.fill = changed_fill
    else:
        print(f"No match found for ALTKN: {post_altkn}")  # Debug print

workbook.save(outputFile)


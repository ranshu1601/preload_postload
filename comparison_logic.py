import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
import os
import shutil

def clean_value(value):
    """Clean and standardize value for comparison"""
    if value is None or pd.isna(value):
        return ''
    # Convert to string, remove all leading/trailing spaces
    cleaned = str(value).strip()
    
    # Handle decimal numbers (convert to integer string)
    try:
        if '.' in cleaned:
            cleaned = str(int(float(cleaned)))
        elif cleaned.isdigit():
            cleaned = str(int(cleaned))
    except ValueError:
        pass
    
    return cleaned

def get_base_column_name(column_name):
    """Get the base column name (first word before any separator or space)"""
    try:
        # Convert to string and uppercase
        col_name = str(column_name).strip().upper()
        
        # Remove all special characters and replace with space
        cleaned_name = ''
        for char in col_name:
            if char.isalnum():
                cleaned_name += char
            else:
                cleaned_name += ' '
        
        # Get the first word
        base_name = cleaned_name.split()[0]
        return base_name
        
    except Exception as e:
        print(f"Error processing column name '{column_name}': {str(e)}")
        return str(column_name).upper()

def get_column_suggestions(pre_columns, post_columns):
    """Get all possible column matches between pre and post data"""
    column_suggestions = {}
    
    print("\nAnalyzing possible column matches:")
    for post_col in post_columns:
        post_base = get_base_column_name(post_col)
        matches = []
        
        for pre_col in pre_columns:
            pre_base = get_base_column_name(pre_col)
            if post_base == pre_base:
                matches.append(pre_col)
                print(f"Found match: {post_col} -> {pre_col} (base: {post_base})")
        
        if matches:
            column_suggestions[post_col] = matches
    
    return column_suggestions

def compare_excel_files(preload_file, postload_file, pre_sheet, post_sheet, key_column, output_dir):
    """Compare two Excel sheets and highlight differences"""
    try:
        # Load Excel sheets
        pre_data = pd.read_excel(preload_file, sheet_name=pre_sheet)
        
        # Create output file path
        output_file = os.path.join(output_dir, 'comparison_result.xlsx')
        
        # If this is the first sheet being processed, copy the postload file
        if not os.path.exists(output_file):
            print("Copying postload file to output file")
            shutil.copy2(postload_file, output_file)
        
        # Load the workbook and get the post-load data
        workbook = load_workbook(output_file)
        post_data = pd.read_excel(postload_file, sheet_name=post_sheet)

        # Find matching columns
        print("\nStarting column mapping process...")
        column_mapping = {}
        for post_col in post_data.columns:
            post_base = get_base_column_name(post_col)
            for pre_col in pre_data.columns:
                if get_base_column_name(pre_col) == post_base:
                    column_mapping[post_col] = pre_col
                    print(f"Matched: {post_col} -> {pre_col}")
                    break

        # Create pre-data dictionary
        pre_dict = {}
        for _, row in pre_data.iterrows():
            key_value = clean_value(row[key_column])
            if key_value:
                pre_dict[key_value] = {col: clean_value(val) for col, val in row.items()}

        # Define colors for highlighting
        changed_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')  # yellow
        missing_fill = PatternFill(start_color='E5A78C', end_color='E5A78C', fill_type='solid')  # red
        blank_key_fill = PatternFill(start_color='E6F3FF', end_color='E6F3FF', fill_type='solid')  # blue

        # Remove the sheet if it exists and create a new one
        if post_sheet in workbook.sheetnames:
            workbook.remove(workbook[post_sheet])
        worksheet = workbook.create_sheet(post_sheet)

        # Write headers
        for col_idx, column_name in enumerate(post_data.columns, 1):
            cell = worksheet.cell(row=1, column=col_idx)
            cell.value = column_name

        # Write data and compare
        for row_idx, row in post_data.iterrows():
            excel_row = row_idx + 2
            
            # Write all values first
            for col_idx, (column_name, value) in enumerate(row.items(), 1):
                cell = worksheet.cell(row=excel_row, column=col_idx)
                cell.value = value

            # Get key value
            key_value = clean_value(row[key_column])
            
            # Handle blank key values
            if not key_value:
                for col_idx in range(1, len(post_data.columns) + 1):
                    worksheet.cell(row=excel_row, column=col_idx).fill = blank_key_fill
                continue

            # Compare values if key exists in pre-data
            if key_value in pre_dict:
                for col_idx, post_col_name in enumerate(post_data.columns, 1):
                    if post_col_name != key_column:  # Skip key column
                        cell = worksheet.cell(row=excel_row, column=col_idx)
                        post_value = clean_value(cell.value)
                        
                        pre_col_name = column_mapping.get(post_col_name)
                        if pre_col_name:
                            pre_value = clean_value(pre_dict[key_value].get(pre_col_name, ''))
                            
                            if post_value != pre_value:
                                if not post_value and pre_value:
                                    cell.fill = missing_fill
                                else:
                                    cell.fill = changed_fill

        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width

        workbook.save(output_file)
        print(f"Comparison completed and saved to {output_file}")
        return output_file

    except Exception as e:
        print(f"Error in compare_excel_files: {str(e)}")
        raise 